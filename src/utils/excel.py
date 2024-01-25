from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Side, Border, PatternFill

from src.entities.scammers.service import scammers_service
from src.entities.scammers.models import Scammer


redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
greenFill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
side = Side(border_style="thin", color="000000")
border = Border(top=side, left=side, right=side, bottom=side)
alignment = Alignment(horizontal="center")


async def create_list_scammer():
    wb = Workbook()
    ws: Worksheet = wb.active
    scammers = await scammers_service.get_scammer_list()

    ws.cell(row=1, column=1, value="№")
    ws.cell(row=1, column=2, value="ID")
    ws.cell(row=1, column=3, value="Username")
    ws.cell(row=1, column=4, value="First Name")
    ws.cell(row=1, column=5, value="Подтверждено")

    index = 2
    scammer: Scammer
    for scammer in scammers:
        ws.cell(row=index, column=1, value=index - 1)
        ws.cell(row=index, column=2, value=scammer.id)
        ws.cell(row=index, column=3, value=scammer.username)
        ws.cell(row=index, column=4, value=scammer.first_name)
        cell = ws.cell(row=index, column=5)
        if scammer.is_scam:
            cell.value = "Да"
            cell.fill = greenFill
        else:
            cell.value = "Не подтверждено"
            cell.fill = redFill
        index += 1

    column_widths = []

    for col_cells in ws.iter_cols(min_col=1, max_col=5):
        cell_lens = []
        cell: Cell
        for cell in col_cells:
            cell.border = border
            cell.alignment = alignment
            cell_lens.append(len(str(cell.value)))
        column_widths.append(max(cell_lens) + 5)

    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = column_widths[i - 1]

    filename = "Список скамеров.xlsx"
    wb.save(filename)
    return filename
